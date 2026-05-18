import argparse
import logging
import sys
import xml.etree.ElementTree as ET

from openpyxl import Workbook

logger = logging.getLogger(__name__)


def return_password_reset_string():
    return "password"


def str2bool(v):
    if v.lower() in ("yes", "true", "t", "y", "1"):
        return True
    elif v.lower() in ("no", "false", "f", "n", "0"):
        return False
    else:
        raise argparse.ArgumentTypeError("Boolean value expected.")


def set_creds(args):
    vtamcredentials = {"user": args.user, "password": args.password}

    if args.appuser is None:
        appcredentials = {"user": args.user}
    else:
        appcredentials = {"user": args.appuser}

    if args.apppassword is None:
        appcredentials["password"] = args.password
    else:
        appcredentials["password"] = args.apppassword

    return {"vtamcredentials": vtamcredentials, "appcredentials": appcredentials}


def do_setup():
    print(r"""


                      ________              __
                     /  _____/______ __ ___/  |_  ____
                    /   \  __\_  __ \  |  \   __\/ __ \
                    \    \_\  \  | \/  |  /|  | \  ___/
                     \______  /__|  |____/ |__|  \___  >
                            \/                       \/
                """)

    parser = argparse.ArgumentParser(
        description="Grute helps with green screen app testing, in theory"
    )

    # Connectivity

    parser.add_argument(
        "-t",
        "--target",
        help="target IP address or Hostname and port: TARGET[:PORT]",
        dest="target",
    )
    parser.add_argument(
        "-s",
        "--sleep",
        help=(
            "Seconds to sleep between actions (increase on slower systems). "
            "The default is 1 second."
        ),
        default=0.5,
        type=float,
        dest="sleep",
    )
    # Config

    parser.add_argument(
        "-v",
        "--visable",
        help="uses x or s 3270.  X is an X window system and is visable\n"
        "Screen is generally used for scripting and goes really fast by "
        "comparision - Bool",
        default=False,
        type=str2bool,
    )
    parser.add_argument(
        "-c",
        "--clobber",
        help="remove target file on run, Bool",
        default=False,
        type=str2bool,
        required=False,
    )
    parser.add_argument(
        "-cfg",
        "--config",
        help="configuration file for application dictionary setup",
        default="default.xml",
    )
    parser.add_argument(
        "-e",
        "--env_switch",
        help="enters a character on the app screen to switch app instance" " - Bool",
        default=False,
        type=str2bool,
    )
    parser.add_argument(
        "-d", "--debug", help="More chatty - bool", default=False, type=str2bool
    )

    # Credentials

    parser.add_argument(
        "-u", "--user", help="supply a username name to connect with", default=None
    )
    parser.add_argument(
        "-p", "--password", help="supply a password name to connect with", default=None
    )
    parser.add_argument(
        "-au",
        "--appuser",
        help="supply a username name to connect with to the app"
        "defaults to the same value as --user",
    )
    parser.add_argument(
        "-ap",
        "--apppassword",
        help="supply a password name to connect with to the app"
        "defaults to the same value as --password",
    )

    # Helpers

    parser.add_argument(
        "-chg", "--changepass", help="Change passwords helper, Bool", type=str2bool
    )
    parser.add_argument(
        "-l",
        "--logmein",
        help="just starts an emulator and logs you in, bool",
        type=str2bool,
        default=False,
    )
    parser.add_argument(
        "-ct",
        "--cemt_trans",
        help="cemt transaction scraping , Bool",
        default=False,
        type=str2bool,
    )

    # MQ

    parser.add_argument("-q", "--que", help="which que do you want", default=False)
    parser.add_argument("-mq", "--mq", help="mq host address", default=False)
    parser.add_argument(
        "-popc",
        "--populate_cics",
        help="populates MQ for cics testing and thats it, Bool",
        type=str2bool,
        default=False,
    )
    parser.add_argument(
        "-popa",
        "--populate_apps",
        help="populates MQ for app testing and thats it, Bool",
        type=str2bool,
        default=False,
    )
    parser.add_argument(
        "-popu",
        "--populate_users",
        help="populates MQ for user testing and thats it, Bool",
        type=str2bool,
        default=False,
    )
    parser.add_argument(
        "-bac",
        "--bulk_auth_create",
        help="creates queues for bulk auth testing - Bool",
        default=False,
        type=str2bool,
    )
    parser.add_argument(
        "-des",
        "--destructive",
        help="Destroys Queues before creation",
        default=False,
        type=str2bool,
    )

    # Types of testing

    parser.add_argument(
        "-cics",
        "--check_cics",
        help="do you want to check cics - bool",
        type=str2bool,
        default=False,
    )
    parser.add_argument(
        "-app",
        "--check_app",
        help="do you want to check cics - bool",
        type=str2bool,
        default=False,
    )
    parser.add_argument(
        "-users",
        "--check_user",
        help="do you want to check userenum - bool",
        type=str2bool,
        default=False,
    )
    parser.add_argument(
        "-ba",
        "--bulk_auth",
        help="bulk app auth testing mode - Bool",
        default=False,
        type=str2bool,
    )
    parser.add_argument(
        "-o", "--overtype", help="overtype testing, Bool", default=False, type=str2bool
    )

    # Spread Sheets

    parser.add_argument(
        "-x",
        "--excel",
        help="everyone just loves spreadsheets - bool",
        default=False,
        type=str2bool,
    )
    parser.add_argument(
        "-xgen",
        "--gen_excel_testing",
        help="Genereates some sample queues for dummy test - bool",
        default=False,
        type=str2bool,
    )

    # MQ Backup and restore

    parser.add_argument(
        "-man",
        "--manual_inport",
        help="do the manual inport thing",
        default=False,
        type=str2bool,
    )
    parser.add_argument(
        "-mane",
        "--manual_export",
        help="do the manual export thing",
        default=False,
        type=str2bool,
    )
    parser.add_argument(
        "-f", "--file_input", help="input file for manual thing", default=False
    )
    parser.add_argument(
        "-fo", "--file_output", help="output file for manual thing", default=False
    )

    # Custom Testing

    parser.add_argument(
        "-dept",
        "--department",
        help="department scraping, Bool",
        default=False,
        type=str2bool,
    )

    # Results DB

    parser.add_argument(
        "--db",
        help="path to SQLite results database (default: phosphor.db)",
        default="phosphor.db",
    )

    args = parser.parse_args()

    if (
        args.populate_cics
        or args.populate_apps
        or args.populate_users
        or args.bulk_auth_create
        or args.excel
    ):
        return args

    if args.manual_export:
        if args.file_output and args.que:
            return args
        else:
            print("[!] You need to specify a queue and backup file. Try -h for help.")
            sys.exit()

    if args.manual_inport:
        if args.file_input and args.que:
            return args
        else:
            print("[!] You need to specify a queue and restore file. Try -h for help.")
            sys.exit()

    if args.logmein:
        args.visable = True

    if not args.target:
        print("      [!] You gotta specify a target. Try -h for help.")
        sys.exit()

    if ":" not in args.target:
        print("      [!] Target Format off, use host:port . Try -h for help.")
        sys.exit()

    return args


def read_xml(configfile, first_level_element_to_find):
    tree = ET.ElementTree(file=configfile)
    found = tree.findall(first_level_element_to_find)
    return [{field.tag: field.text for field in instance} for instance in found]


def make_excel_sheets(wb, user_list_dict, env_name):
    ws = wb.create_sheet(env_name, 0)  # insert at first position
    ws.sheet_properties.tabColor = "1072BA"
    ws.merge_cells("B1:E1")
    ws["A1"] = "Transaction"
    ws["B1"] = "Users"
    col = 2
    for user_dict in user_list_dict:
        ws.cell(row=2, column=col, value=user_dict["user"])
        col += 1

    return wb


def save_excel_workbook(wb, fn="Example.xlsx"):
    wb.save(fn)


def make_excel_workbook(user_list_dict, env_list_dict):
    wb = Workbook()

    for env_dictionary in env_list_dict:
        wb = make_excel_sheets(wb, user_list_dict, env_dictionary["name"])

    return wb


def search_for_previous_application_code(ws, application_code):
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is not None:
            if ws.cell(row=row, column=1).value == application_code.lower():
                return row

    return None


def search_for_user_column(ws, user):
    for rowOfCellObjects in ws["B2":"E2"]:
        for cellObj in rowOfCellObjects:
            if cellObj.value == user:
                return cellObj.column

    logger.warning("No user match for %s", user)
    return False


def process_mq_results_into_excel(
    wb, user, enviroment, response_type, application_code
):
    ws = wb[enviroment]

    user_column = search_for_user_column(ws, user)

    if not user_column:
        logger.error("No user column found for %s, cannot write result", user)
        sys.exit()

    pre_existing_application_code = search_for_previous_application_code(
        ws, application_code
    )

    if pre_existing_application_code is None:
        ws.append({"A": application_code, user_column: response_type})
    else:
        ws.cell(
            column=user_column, row=pre_existing_application_code, value=response_type
        )

    return wb
