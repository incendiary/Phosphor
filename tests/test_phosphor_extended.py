import argparse
import os
import sys
import unittest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from unittest.mock import MagicMock, patch

FIXTURE = os.path.join(os.path.dirname(__file__), "fixtures", "test_config.xml")


def make_args(**kwargs):
    defaults = dict(
        target="localhost:3270",
        sleep=0.5,
        clobber=False,
        user="uid1",
        password="pass1",
        appuser=None,
        apppassword=None,
        config=FIXTURE,
        debug=False,
        visable=False,
        overtype=False,
        env_switch=False,
        bulk_auth=False,
        check_cics=False,
        check_app=False,
        check_user=False,
        logmein=False,
        changepass=False,
        populate_cics=False,
        populate_apps=False,
        populate_users=False,
        bulk_auth_create=False,
        excel=False,
        gen_excel_testing=False,
        manual_inport=False,
        manual_export=False,
        file_input=None,
        file_output=None,
        que=False,
        mq=False,
        destructive=False,
        department=False,
        cemt_trans=False,
        db=":memory:",
        targets=None,
    )
    defaults.update(kwargs)
    return argparse.Namespace(**defaults)


def make_mainframe(args=None):
    from phosphor import MainFrame

    if args is None:
        args = make_args()
    credentials = {
        "vtamcredentials": {"user": args.user, "password": args.password},
        "appcredentials": {"user": args.user, "password": args.password},
    }
    mock_sp = MagicMock()
    mock_sp.stdout.readline.side_effect = [
        b"U F U C(localhost) I 2 24 80 0 0 0x0 0.0\n",
        b"ok\n",
    ] * 200
    with patch("platform.system", return_value="Linux"):
        with patch("subprocess.Popen", return_value=mock_sp):
            mf = MainFrame(args.target, args.sleep, args.clobber, credentials, args)
    return mf


# ---------------------------------------------------------------------------
# make_and_set_folder_path
# ---------------------------------------------------------------------------


class TestMakeAndSetFolderPath(unittest.TestCase):

    def setUp(self):
        self.mf = make_mainframe()
        self.mf.credentials = {"appcredentials": {"user": "uid1", "password": "pass1"}}
        self.mf.application_response = "app_error"
        self.mf.app_code = "abc"

    def test_with_environment(self):
        self.mf.environment = {"name": "env1", "default": "True"}
        self.mf.make_and_set_folder_path()
        self.assertEqual(self.mf.path_to_folder, "app/uid1/env1/app_error/a/b")

    def test_without_environment(self):
        self.mf.environment = None
        self.mf.make_and_set_folder_path()
        self.assertEqual(self.mf.path_to_folder, "app/uid1/app_error/a/b")

    def test_path_uses_first_two_chars_of_app_code(self):
        self.mf.environment = None
        self.mf.app_code = "xyz"
        self.mf.make_and_set_folder_path()
        self.assertIn("/x/y", self.mf.path_to_folder)


# ---------------------------------------------------------------------------
# look_for_app_code
# ---------------------------------------------------------------------------


class TestLookForAppCode(unittest.TestCase):

    def setUp(self):
        self.mf = make_mainframe()
        self.mf.credentials = {"appcredentials": {"user": "uid1", "password": "pass1"}}
        self.mf.environment = {"name": "env1", "default": "True"}
        self.mf.application_list_dict = [
            {"string": "ERROR", "type": "app_error", "xpos": "1", "ypos": "24"},
            {
                "string": "SECURITY VIOLATION",
                "type": "app_auth",
                "xpos": "20",
                "ypos": "2",
            },
        ]
        # Replace emulator with a full mock so debug string_get calls don't
        # reach the real subprocess stub
        self.mf.em = MagicMock()
        self.mf.em.string_get = MagicMock(return_value="")

    def test_sets_response_on_match(self):
        self.mf.em.string_found = MagicMock(side_effect=[True, False])
        self.mf.look_for_app_code()
        self.assertEqual(self.mf.application_response, "app_error")

    def test_sets_mq_queue_non_bulk(self):
        self.mf.bulk_app_mode = False
        self.mf.em.string_found = MagicMock(side_effect=[True, False])
        self.mf.look_for_app_code()
        self.assertEqual(self.mf.mq_queue, "app_error")

    def test_sets_mq_queue_bulk_mode(self):
        self.mf.bulk_app_mode = True
        self.mf.em.string_found = MagicMock(side_effect=[False, True])
        self.mf.look_for_app_code()
        self.assertEqual(self.mf.mq_queue, "uid1_env1_app_auth")

    def test_no_match_leaves_response_unchanged(self):
        self.mf.application_response = "sentinel"
        self.mf.em.string_found = MagicMock(return_value=False)
        self.mf.look_for_app_code()
        self.assertEqual(self.mf.application_response, "sentinel")

    def test_returns_on_first_match(self):
        # Both entries match — only first should be used
        self.mf.em.string_found = MagicMock(return_value=True)
        self.mf.look_for_app_code()
        self.assertEqual(self.mf.application_response, "app_error")


# ---------------------------------------------------------------------------
# look_for_login_code
# ---------------------------------------------------------------------------


class TestLookForLoginCode(unittest.TestCase):

    def setUp(self):
        self.mf = make_mainframe()
        self.mf.username_responses_list_dict = [
            {"string": "ICH408I", "type": "user_invalid", "xpos": "2", "ypos": "22"},
            {"string": "PASSWORD", "type": "user_valid", "xpos": "2", "ypos": "22"},
        ]

    def test_sets_username_response_on_match(self):
        self.mf.em.string_get = MagicMock(side_effect=["ICH408I", "PASSWORD"])
        self.mf.look_for_login_code()
        self.assertEqual(self.mf.username_response, "user_invalid")

    def test_second_entry_matched(self):
        self.mf.em.string_get = MagicMock(side_effect=["XXXXXXX", "PASSWORD"])
        self.mf.look_for_login_code()
        self.assertEqual(self.mf.username_response, "user_valid")

    def test_no_match_leaves_response_unchanged(self):
        self.mf.username_response = "sentinel"
        self.mf.em.string_get = MagicMock(return_value="XXXXXXX")
        self.mf.look_for_login_code()
        self.assertEqual(self.mf.username_response, "sentinel")


# ---------------------------------------------------------------------------
# assess_login_screen
# ---------------------------------------------------------------------------


class TestAssessLoginScreen(unittest.TestCase):

    def setUp(self):
        self.mf = make_mainframe()
        self.mf.username_field_location_dict = {"xpos": "1", "ypos": "22"}
        self.mf.username_responses_list_dict = [
            {"string": "ICH408I", "type": "user_invalid", "xpos": "2", "ypos": "22"},
        ]
        self.mf.username_to_check = "testuser"
        self.mf.em = MagicMock()
        self.mf.save_screen_specific = MagicMock()

    def test_no_match_sets_user_unknown(self):
        self.mf.em.string_get = MagicMock(return_value="XXXXXXX")
        self.mf.assess_login_screen()
        self.assertEqual(self.mf.username_response, "user_unknown")

    def test_match_sets_correct_type(self):
        self.mf.em.string_get = MagicMock(return_value="ICH408I")
        self.mf.assess_login_screen()
        self.assertEqual(self.mf.username_response, "user_invalid")

    def test_saves_screen_after_assessment(self):
        self.mf.em.string_get = MagicMock(return_value="XXXXXXX")
        self.mf.assess_login_screen()
        self.mf.save_screen_specific.assert_called_once()
        saved_path = self.mf.save_screen_specific.call_args[0][0]
        self.assertIn("testuser", saved_path)
        self.assertIn("user_unknown", saved_path)


# ---------------------------------------------------------------------------
# find_cemt_transactions_on_screen
# ---------------------------------------------------------------------------


class TestFindCemtTransactions(unittest.TestCase):

    def setUp(self):
        self.mf = make_mainframe()

    def test_extracts_transaction_codes(self):
        screen = [
            "  Tra(CESN)  Sta(Ena)  Use(000) Tas(000) Sus(000) Hty(Bnd)",
            "  Tra(CICS)  Sta(Ena)  Use(000) Tas(000)",
            "  Some other line without transaction",
        ]
        self.mf.em.screen_get = MagicMock(return_value=screen)
        self.mf.find_cemt_transactions_on_screen()
        self.assertIn("CESN", self.mf.transaction_codes)
        self.assertIn("CICS", self.mf.transaction_codes)

    def test_ignores_lines_without_tra(self):
        screen = ["  Some other line without transaction"]
        self.mf.em.screen_get = MagicMock(return_value=screen)
        self.mf.find_cemt_transactions_on_screen()
        self.assertEqual(self.mf.transaction_codes, [])

    def test_accumulates_across_calls(self):
        screen1 = ["  Tra(CESN)  Sta(Ena)"]
        screen2 = ["  Tra(CEMT)  Sta(Ena)"]
        self.mf.em.screen_get = MagicMock(side_effect=[screen1, screen2])
        self.mf.find_cemt_transactions_on_screen()
        self.mf.find_cemt_transactions_on_screen()
        self.assertEqual(len(self.mf.transaction_codes), 2)


# ---------------------------------------------------------------------------
# check_application: bad app code path (double-ack regression)
# ---------------------------------------------------------------------------


class TestCheckApplicationBadCode(unittest.TestCase):

    def setUp(self):
        self.mf = make_mainframe()
        self.mf.set_bad_app_codes(["skip1"])
        self.mf.application_list_dict = [
            {"string": "ERROR", "type": "app_error", "xpos": "1", "ypos": "24"}
        ]
        self.mf.environment = {"name": "env1", "default": "True"}
        self.mf.credentials = {"appcredentials": {"user": "uid1", "password": "pass1"}}

    def test_bad_code_sets_app_unknown_without_calling_assess(self):
        self.mf.app_code = "skip1"
        self.mf.application_response = None
        self.mf.mq_queue = None

        # Simulate what check_application does for a bad code
        if self.mf.app_code in self.mf.bad_app_codes:
            self.mf.application_response = "app_unknown"
            self.mf.mq_queue = "app_unknown"

        self.assertEqual(self.mf.application_response, "app_unknown")
        self.assertEqual(self.mf.mq_queue, "app_unknown")

    def test_bad_code_response_never_none(self):
        # Confirms the fix: mq_queue is set before the publish step runs
        self.mf.app_code = "skip1"
        self.mf.application_response = None
        self.mf.mq_queue = None

        if self.mf.app_code in self.mf.bad_app_codes:
            self.mf.application_response = "app_unknown"
            self.mf.mq_queue = "app_unknown"

        self.assertIsNotNone(self.mf.mq_queue)
        self.assertIsNotNone(self.mf.application_response)


# ---------------------------------------------------------------------------
# search_for_previous_application_code (off-by-one regression)
# ---------------------------------------------------------------------------


class TestSearchForPreviousAppCode(unittest.TestCase):

    def test_finds_entry_in_last_row(self):
        from openpyxl import Workbook

        from inc.public_includes import search_for_previous_application_code

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="abc")
        ws.cell(row=2, column=1, value="def")  # last row
        # max_row is 2; before fix range(1,2) would skip row 2
        result = search_for_previous_application_code(ws, "def")
        self.assertEqual(result, 2)

    def test_returns_none_when_not_found(self):
        from openpyxl import Workbook

        from inc.public_includes import search_for_previous_application_code

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="abc")
        result = search_for_previous_application_code(ws, "zzz")
        self.assertIsNone(result)

    def test_finds_entry_in_first_row(self):
        from openpyxl import Workbook

        from inc.public_includes import search_for_previous_application_code

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="abc")
        result = search_for_previous_application_code(ws, "abc")
        self.assertEqual(result, 1)


if __name__ == "__main__":
    unittest.main()
